import os

from celery import shared_task
from django.conf import settings
from django.core.files import File
from openpyxl import Workbook, load_workbook
from django.conf import settings

from .models import Widget


@shared_task
def add(x, y):
    return x + y


@shared_task
def mul(x, y):
    return x * y


@shared_task
def xsum(numbers):
    return sum(numbers)


@shared_task
def count_widgets():
    return Widget.objects.count()


@shared_task
def rename_widget(widget_id, name):
    w = Widget.objects.get(id=widget_id)
    w.name = name
    w.save()


@shared_task
def do_osinfo_query(input_file):
    wb = load_workbook(os.path.join(
        settings.MEDIA_ROOT, 'osinfo', 'input', input_file))
    ws = wb.active
    colA = ws['A']
    for cell in colA:
        print(cell.value)
    # import time
    # from datetime import datetime
    # wb = Workbook()
    # # do some stuff
    # time.sleep(1)
    # file_name = f"out_put_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    # file_path = os.path.join(
    #     settings.MEDIA_ROOT, 'osinfo', 'output', file_name)
    # wb.save(file_path)

    # with open(file_path, 'rb') as f:
    #     myfile = File(f)
